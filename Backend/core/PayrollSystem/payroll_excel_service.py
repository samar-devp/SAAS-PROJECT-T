"""
Payroll Excel Export/Import Service
Handles Excel operations for payroll data
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from io import BytesIO
from django.http import HttpResponse
from decimal import Decimal
import uuid


class PayrollExcelService:
    """Service for payroll Excel operations"""
    
    @staticmethod
    def generate_payroll_excel(payroll_records, month, year):
        """Generate comprehensive payroll Excel report"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Payroll_{month}_{year}"
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            "Employee ID", "Employee Name", "Email",
            "Basic", "HRA", "Special Allowance", "Overtime", "Gross Salary",
            "PF (Emp)", "ESI (Emp)", "Professional Tax", "TDS", "Advance", "Loan",
            "Total Deductions", "Net Salary",
            "Present Days", "Absent Days", "Working Days"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Write data
        for row_idx, payroll in enumerate(payroll_records, 2):
            try:
                user_profile = payroll.employee.own_user_profile
                row_data = [
                    user_profile.custom_employee_id,
                    user_profile.user_name,
                    payroll.employee.email,
                    float(payroll.basic_salary),
                    float(payroll.hra),
                    float(payroll.special_allowance),
                    float(payroll.overtime_amount),
                    float(payroll.gross_salary),
                    float(payroll.pf_employee),
                    float(payroll.esi_employee),
                    float(payroll.professional_tax),
                    float(payroll.tds),
                    float(payroll.advance_deduction),
                    float(payroll.loan_deduction),
                    float(payroll.total_deductions),
                    float(payroll.net_salary),
                    payroll.present_days,
                    payroll.absent_days,
                    payroll.working_days
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = border
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        cell.alignment = Alignment(horizontal="left")
            except Exception as e:
                continue
        
        # Auto width
        for col in ws.columns:
            max_len = 0
            letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[letter].width = min(max_len + 2, 30)
        
        # Save to memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="payroll_{month}_{year}.xlsx"'
        
        return response
    
    @staticmethod
    def generate_sample_excel(component_type='deduction'):
        """Generate sample Excel for upload"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Sample_{component_type.title()}"
        
        headers = ["Employee ID", "Component Code", "Amount", "Remarks"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Sample data row
        sample_data = ["EMP001", "ADVANCE", "5000", "Sample advance deduction"]
        for col, value in enumerate(sample_data, 1):
            ws.cell(row=2, column=col, value=value)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="sample_{component_type}.xlsx"'
        
        return response

