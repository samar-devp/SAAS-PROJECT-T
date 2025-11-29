"""
Additional Payroll Views - Excel, Reports, Payslips
"""

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.db.models import Q, Sum, Count
from datetime import date
from decimal import Decimal
import openpyxl
from io import BytesIO
from django.http import HttpResponse

from .models import PayrollRecord, SalaryComponent, PayrollComponentEntry
from .serializers import PayrollRecordSerializer
from .payroll_excel_service import PayrollExcelService
from AuthN.models import BaseUserModel, UserProfile
from utils.pagination_utils import CustomPagination


# ==================== UPDATE PAYROLL ====================

class UpdatePayrollReportView(APIView):
    """Update existing payroll record"""
    permission_classes = [IsAuthenticated]
    
    @transaction.atomic
    def put(self, request, clnID, uid):
        """Update payroll for specific employee"""
        try:
            from .payroll_calculator import PayrollCalculator
            
            organization = get_object_or_404(BaseUserModel, id=clnID, role='organization')
            employee = get_object_or_404(BaseUserModel, id=uid, role='user')
            
            month = int(request.data.get('month'))
            year = int(request.data.get('year'))
            
            payroll = get_object_or_404(
                PayrollRecord,
                employee=employee,
                organization=organization,
                payroll_month=month,
                payroll_year=year
            )
            
            # Recalculate payroll
            user_profile = employee.own_user_profile
            calculator = PayrollCalculator(employee, month, year, user_profile.admin, organization)
            payroll_data = calculator.calculate_payroll()
            
            # Update payroll record
            payroll.basic_salary = payroll_data['earnings'].get('BASIC', {}).get('amount', 0)
            payroll.hra = payroll_data['earnings'].get('HRA', {}).get('amount', 0)
            payroll.special_allowance = payroll_data['earnings'].get('SPECIAL_ALLOWANCE', {}).get('amount', 0)
            payroll.overtime_amount = payroll_data['earnings'].get('OVERTIME', {}).get('amount', 0)
            payroll.gross_salary = payroll_data['gross_salary']
            payroll.pf_employee = payroll_data['deductions'].get('PF_EMPLOYEE', {}).get('amount', 0)
            payroll.esi_employee = payroll_data['deductions'].get('ESI_EMPLOYEE', {}).get('amount', 0)
            payroll.professional_tax = payroll_data['deductions'].get('PROFESSIONAL_TAX', {}).get('amount', 0)
            payroll.tds = payroll_data['deductions'].get('TDS', {}).get('amount', 0)
            payroll.advance_deduction = payroll_data['deductions'].get('ADVANCE', {}).get('amount', 0)
            payroll.loan_deduction = payroll_data['deductions'].get('LOAN', {}).get('amount', 0)
            payroll.total_deductions = payroll_data['total_deductions']
            payroll.net_salary = payroll_data['net_salary']
            payroll.earnings_breakdown = payroll_data['earnings']
            payroll.deductions_breakdown = payroll_data['deductions']
            payroll.save()
            
            # Update component entries
            PayrollComponentEntry.objects.filter(payroll=payroll).delete()
            
            for code, comp_data in payroll_data['earnings'].items():
                try:
                    component = SalaryComponent.objects.get(code=code, organization=organization)
                    PayrollComponentEntry.objects.create(
                        payroll=payroll,
                        component=component,
                        amount=comp_data['amount'],
                        is_earning=True
                    )
                except SalaryComponent.DoesNotExist:
                    pass
            
            for code, comp_data in payroll_data['deductions'].items():
                try:
                    component = SalaryComponent.objects.get(code=code, organization=organization)
                    PayrollComponentEntry.objects.create(
                        payroll=payroll,
                        component=component,
                        amount=comp_data['amount'],
                        is_earning=False
                    )
                except SalaryComponent.DoesNotExist:
                    pass
            
            serializer = PayrollRecordSerializer(payroll)
            return Response({
                "status": status.HTTP_200_OK,
                "message": "Payroll updated successfully",
                "data": serializer.data
            })
            
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==================== BI PAYROLL REPORT ====================

class BIPayrollReportView(APIView):
    """Business Intelligence Payroll Report with filters"""
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination
    
    def get(self, request, OrgID, month, year, clnID=None, uid=None):
        """Get BI payroll report with various filters"""
        try:
            organization = get_object_or_404(BaseUserModel, id=OrgID, role='organization')
            
            queryset = PayrollRecord.objects.filter(
                organization=organization,
                payroll_month=month,
                payroll_year=year
            )
            
            if clnID:
                admin = get_object_or_404(BaseUserModel, id=clnID, role='admin')
                queryset = queryset.filter(admin=admin)
            
            if uid:
                employee = get_object_or_404(BaseUserModel, id=uid, role='user')
                queryset = queryset.filter(employee=employee)
            
            # Pagination
            paginator = self.pagination_class()
            paginated_qs = paginator.paginate_queryset(queryset, request)
            
            serializer = PayrollRecordSerializer(paginated_qs, many=True)
            pagination_data = paginator.get_paginated_response(serializer.data)
            pagination_data["results"] = serializer.data
            
            # Detailed Summary
            summary = queryset.aggregate(
                total_employees=Count('id'),
                total_gross=Sum('gross_salary'),
                total_deductions=Sum('total_deductions'),
                total_net=Sum('net_salary'),
                total_pf_employee=Sum('pf_employee'),
                total_pf_employer=Sum('pf_employer'),
                total_esi_employee=Sum('esi_employee'),
                total_esi_employer=Sum('esi_employer'),
                total_tds=Sum('tds'),
                total_professional_tax=Sum('professional_tax'),
                total_advance=Sum('advance_deduction'),
                total_loan=Sum('loan_deduction')
            )
            
            # Calculate averages
            count = queryset.count()
            if count > 0:
                summary['avg_gross'] = summary['total_gross'] / count if summary['total_gross'] else 0
                summary['avg_net'] = summary['total_net'] / count if summary['total_net'] else 0
            else:
                summary['avg_gross'] = 0
                summary['avg_net'] = 0
            
            pagination_data["summary"] = summary
            pagination_data["message"] = "BI Payroll report fetched successfully"
            
            return Response(pagination_data)
            
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==================== PAYROLL DOWNLOAD INFO ====================

class PayrollDownloadInfo(APIView):
    """Get payroll download information for Excel export"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request, org_id, month, year):
        """Get payroll data formatted for download"""
        try:
            organization = get_object_or_404(BaseUserModel, id=org_id, role='organization')
            
            payrolls = PayrollRecord.objects.filter(
                organization=organization,
                payroll_month=month,
                payroll_year=year
            ).select_related('employee', 'admin')
            
            download_data = []
            for payroll in payrolls:
                try:
                    user_profile = payroll.employee.own_user_profile
                    download_data.append({
                        'employee_id': user_profile.custom_employee_id,
                        'employee_name': user_profile.user_name,
                        'email': payroll.employee.email,
                        'basic_salary': float(payroll.basic_salary),
                        'hra': float(payroll.hra),
                        'special_allowance': float(payroll.special_allowance),
                        'overtime': float(payroll.overtime_amount),
                        'gross_salary': float(payroll.gross_salary),
                        'pf_employee': float(payroll.pf_employee),
                        'esi_employee': float(payroll.esi_employee),
                        'professional_tax': float(payroll.professional_tax),
                        'tds': float(payroll.tds),
                        'advance': float(payroll.advance_deduction),
                        'loan': float(payroll.loan_deduction),
                        'total_deductions': float(payroll.total_deductions),
                        'net_salary': float(payroll.net_salary),
                        'present_days': payroll.present_days,
                        'absent_days': payroll.absent_days,
                        'working_days': payroll.working_days
                    })
                except:
                    continue
            
            return Response({
                "status": status.HTTP_200_OK,
                "message": "Download info fetched successfully",
                "data": download_data
            })
            
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==================== EXCEL EXPORT/IMPORT ====================

class DownloadAdvanceIncomeTaxSampleExcel(APIView):
    """Download sample Excel for advance/income tax upload"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """Download sample Excel"""
        return PayrollExcelService.generate_sample_excel('deduction')


class DownloadDeductionSampleExcel(APIView):
    """Download sample Excel for deductions"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """Download sample Excel"""
        return PayrollExcelService.generate_sample_excel('deduction')


class DownloadEarningSampleExcel(APIView):
    """Download sample Excel for earnings"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """Download sample Excel"""
        return PayrollExcelService.generate_sample_excel('earning')


class UploadAdvanceIncomeTaxExcel(APIView):
    """Upload Excel for advance/income tax"""
    permission_classes = [IsAuthenticated]
    
    def post(self, request, clnID, month, year):
        """Upload and process Excel file"""
        try:
            organization = get_object_or_404(BaseUserModel, id=clnID, role='organization')
            
            if 'file' not in request.FILES:
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "No file uploaded"
                }, status=status.HTTP_400_BAD_REQUEST)
            
            excel_file = request.FILES['file']
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            processed = 0
            errors = []
            
            # Skip header row
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    if not row[0] or not row[1]:
                        continue
                        
                    employee_id = str(row[0]).strip()
                    component_code = str(row[1]).strip()
                    amount = Decimal(str(row[2] or 0))
                    
                    # Find employee
                    user_profile = UserProfile.objects.filter(
                        custom_employee_id=employee_id,
                        organization=organization
                    ).first()
                    
                    if not user_profile:
                        errors.append(f"Employee not found: {employee_id}")
                        continue
                    
                    employee = user_profile.user
                    
                    # Find component
                    component = SalaryComponent.objects.filter(
                        code=component_code,
                        organization=organization,
                        component_type='deduction'
                    ).first()
                    
                    if not component:
                        errors.append(f"Component not found: {component_code}")
                        continue
                    
                    # Get or create payroll record
                    payroll, created = PayrollRecord.objects.get_or_create(
                        employee=employee,
                        organization=organization,
                        payroll_month=month,
                        payroll_year=year,
                        defaults={
                            'admin': user_profile.admin,
                            'payroll_date': date(year, month, 1),
                            'status': 'draft'
                        }
                    )
                    
                    # Update deduction
                    if payroll.deductions_breakdown is None:
                        payroll.deductions_breakdown = {}
                    
                    payroll.deductions_breakdown[component_code] = {
                        'name': component.name,
                        'amount': float(amount)
                    }
                    
                    # Recalculate total deductions
                    total_deductions = sum([
                        Decimal(str(v.get('amount', 0)))
                        for v in payroll.deductions_breakdown.values()
                    ])
                    payroll.total_deductions = total_deductions
                    payroll.net_salary = payroll.gross_salary - total_deductions
                    payroll.save()
                    
                    processed += 1
                    
                except Exception as e:
                    errors.append(f"Row error: {str(e)}")
                    continue
            
            return Response({
                "status": status.HTTP_200_OK,
                "message": f"Processed {processed} records",
                "processed": processed,
                "errors": errors if errors else None
            })
            
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class UploadDeductionExcel(APIView):
    """Upload Excel for deductions"""
    permission_classes = [IsAuthenticated]
    
    def post(self, request, clnID, month, year):
        """Upload deductions Excel"""
        return UploadAdvanceIncomeTaxExcel().post(request, clnID, month, year)


class UploadEarningExcel(APIView):
    """Upload Excel for earnings"""
    permission_classes = [IsAuthenticated]
    
    def post(self, request, clnID, month, year):
        """Upload earnings Excel"""
        try:
            organization = get_object_or_404(BaseUserModel, id=clnID, role='organization')
            
            if 'file' not in request.FILES:
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "No file uploaded"
                }, status=status.HTTP_400_BAD_REQUEST)
            
            excel_file = request.FILES['file']
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            processed = 0
            errors = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    if not row[0] or not row[1]:
                        continue
                        
                    employee_id = str(row[0]).strip()
                    component_code = str(row[1]).strip()
                    amount = Decimal(str(row[2] or 0))
                    
                    user_profile = UserProfile.objects.filter(
                        custom_employee_id=employee_id,
                        organization=organization
                    ).first()
                    
                    if not user_profile:
                        errors.append(f"Employee not found: {employee_id}")
                        continue
                    
                    employee = user_profile.user
                    
                    component = SalaryComponent.objects.filter(
                        code=component_code,
                        organization=organization,
                        component_type='earning'
                    ).first()
                    
                    if not component:
                        errors.append(f"Component not found: {component_code}")
                        continue
                    
                    payroll, created = PayrollRecord.objects.get_or_create(
                        employee=employee,
                        organization=organization,
                        payroll_month=month,
                        payroll_year=year,
                        defaults={
                            'admin': user_profile.admin,
                            'payroll_date': date(year, month, 1),
                            'status': 'draft'
                        }
                    )
                    
                    if payroll.earnings_breakdown is None:
                        payroll.earnings_breakdown = {}
                    
                    payroll.earnings_breakdown[component_code] = {
                        'name': component.name,
                        'amount': float(amount)
                    }
                    
                    # Recalculate gross
                    total_earnings = sum([
                        Decimal(str(v.get('amount', 0)))
                        for v in payroll.earnings_breakdown.values()
                    ])
                    payroll.gross_salary = total_earnings
                    payroll.net_salary = payroll.gross_salary - payroll.total_deductions
                    payroll.save()
                    
                    processed += 1
                    
                except Exception as e:
                    errors.append(f"Row error: {str(e)}")
                    continue
            
            return Response({
                "status": status.HTTP_200_OK,
                "message": f"Processed {processed} records",
                "processed": processed,
                "errors": errors if errors else None
            })
            
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==================== CUSTOM PAYROLL SHEET ====================

class GenerateCustomPayableSheet(APIView):
    """Generate custom payroll sheet"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request, org_id, admin_id, month=None, year=None):
        """Generate custom payroll Excel"""
        try:
            organization = get_object_or_404(BaseUserModel, id=org_id, role='organization')
            admin = get_object_or_404(BaseUserModel, id=admin_id, role='admin')
            
            if month and year:
                # Generate for specific month
                payrolls = PayrollRecord.objects.filter(
                    organization=organization,
                    admin=admin,
                    payroll_month=month,
                    payroll_year=year
                )
                return PayrollExcelService.generate_payroll_excel(payrolls, month, year)
            else:
                # Return settings or template
                return Response({
                    "status": status.HTTP_200_OK,
                    "message": "Provide month and year for payroll generation"
                })
                
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==================== EMPLOYEE PAYSLIP ====================

class EmployeePayslipAPI(APIView):
    """Employee payslip generation and retrieval"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request, admin_id, month, year, uniqueID=None):
        """Get payslip(s)"""
        try:
            admin = get_object_or_404(BaseUserModel, id=admin_id, role='admin')
            
            if uniqueID:
                # Get specific employee payslip
                user_profile = UserProfile.objects.filter(
                    custom_employee_id=uniqueID,
                    admin=admin
                ).first()
                
                if not user_profile:
                    return Response({
                        "status": status.HTTP_404_NOT_FOUND,
                        "message": "Employee not found",
                        "data": []
                    }, status=status.HTTP_404_NOT_FOUND)
                
                payroll = PayrollRecord.objects.filter(
                    employee=user_profile.user,
                    admin=admin,
                    payroll_month=month,
                    payroll_year=year
                ).first()
                
                if not payroll:
                    return Response({
                        "status": status.HTTP_404_NOT_FOUND,
                        "message": "Payroll not found",
                        "data": []
                    }, status=status.HTTP_404_NOT_FOUND)
                
                serializer = PayrollRecordSerializer(payroll)
                return Response({
                    "status": status.HTTP_200_OK,
                    "message": "Payslip fetched successfully",
                    "data": [serializer.data]
                })
            else:
                # Get all employees under admin
                user_profiles = UserProfile.objects.filter(admin=admin)
                employee_ids = [profile.user.id for profile in user_profiles]
                
                payrolls = PayrollRecord.objects.filter(
                    employee_id__in=employee_ids,
                    admin=admin,
                    payroll_month=month,
                    payroll_year=year
                )
                
                serializer = PayrollRecordSerializer(payrolls, many=True)
                return Response({
                    "status": status.HTTP_200_OK,
                    "message": "Payslips fetched successfully",
                    "data": serializer.data
                })
                
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

