import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from io import BytesIO
from django.http import HttpResponse
import uuid


class ExcelExportService:

    @staticmethod
    def generate(attendance_list, attendance_date):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        headers = [
            "Employee Name", "Employee ID", "Email",
            "Status", "Last Login", "Check In", "Check Out",
            "Break (minutes)", "Late (minutes)", "Production (minutes)"
        ]

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Header Row
        for col, head in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=head)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center")

        # Helper function to convert values to Excel-compatible format
        def to_excel_value(val):
            """Convert value to Excel-compatible format"""
            if val is None:
                return "N/A"
            # Convert UUID to string
            if isinstance(val, uuid.UUID):
                return str(val)
            # Check if it's a UUID-like object
            if hasattr(val, '__class__') and 'UUID' in str(type(val)):
                return str(val)
            # Convert datetime objects to string
            if hasattr(val, 'strftime'):
                return val.strftime('%Y-%m-%d %H:%M:%S')
            # Convert any other non-serializable objects to string
            try:
                # Try to convert to string if it's not a basic type
                if not isinstance(val, (str, int, float, bool)):
                    return str(val)
            except:
                return "N/A"
            return val

        # Data Rows
        for i, item in enumerate(attendance_list, 2):
            row = [
                to_excel_value(item.get("employee_name", "N/A")),
                to_excel_value(item.get("employee_id", "N/A")),
                to_excel_value(item.get("employee_email", "N/A")),
                to_excel_value(item.get("attendance_status", "N/A")),
                to_excel_value(item.get("last_login_status") or "N/A"),
                to_excel_value(item.get("check_in") or "N/A"),
                to_excel_value(item.get("check_out") or "N/A"),
                item.get("total_break_minutes", 0) or 0,  # Use raw minutes instead of formatted string
                item.get("late_minutes", 0) or 0,
                item.get("total_working_minutes", 0) or 0,  # Fixed: was production_minutes
            ]
            for col, val in enumerate(row, 1):
                ws.cell(row=i, column=col).value = to_excel_value(val)

        # Auto width
        for col in ws.columns:
            max_len = 0
            letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[letter].width = max_len + 2

        # Save in memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="attendance_{attendance_date}.xlsx"'

        return response
