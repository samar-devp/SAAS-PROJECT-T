"""
Leave Management Excel Export/Import Service
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from io import BytesIO
from django.http import HttpResponse
from decimal import Decimal
from datetime import date


class LeaveExcelService:
    """Service for leave Excel operations"""
    
    @staticmethod
    def generate_leave_balance_excel(leave_balances, year):
        """Generate leave balance Excel report"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Leave_Balance_{year}"
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            "Employee ID", "Employee Name", "Email",
            "Leave Type", "Assigned", "Used", "Pending", "Balance",
            "Carried Forward", "Encashed", "Lapsed"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Write data
        for row_idx, balance in enumerate(leave_balances, 2):
            try:
                user_profile = balance.user.own_user_profile
                row_data = [
                    user_profile.custom_employee_id,
                    user_profile.user_name,
                    balance.user.email,
                    balance.leave_type.code,
                    float(balance.assigned),
                    float(balance.used),
                    float(balance.pending),
                    float(balance.balance),
                    float(balance.carried_forward),
                    float(balance.encashed),
                    float(balance.lapsed)
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = border
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        cell.alignment = Alignment(horizontal="left")
            except Exception as e:
                continue
        
        # Auto width
        for col in ws.columns:
            max_len = 0
            letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[letter].width = min(max_len + 2, 30)
        
        # Save to memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="leave_balance_{year}.xlsx"'
        
        return response
    
    @staticmethod
    def generate_leave_application_excel(leave_applications, month=None, year=None):
        """Generate leave application Excel report"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Leave_Applications"
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            "Employee ID", "Employee Name", "Email",
            "Leave Type", "From Date", "To Date", "Total Days",
            "Status", "Applied At", "Reviewed By", "Comments"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Write data
        for row_idx, application in enumerate(leave_applications, 2):
            try:
                user_profile = application.user.own_user_profile
                row_data = [
                    user_profile.custom_employee_id,
                    user_profile.user_name,
                    application.user.email,
                    application.leave_type.code,
                    application.from_date.isoformat(),
                    application.to_date.isoformat(),
                    float(application.total_days),
                    application.status,
                    application.applied_at.strftime('%Y-%m-%d %H:%M:%S') if application.applied_at else '',
                    application.reviewed_by.email if application.reviewed_by else '',
                    application.comments or ''
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = border
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        cell.alignment = Alignment(horizontal="left")
            except Exception as e:
                continue
        
        # Auto width
        for col in ws.columns:
            max_len = 0
            letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[letter].width = min(max_len + 2, 30)
        
        # Save to memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"leave_applications_{month}_{year}.xlsx" if month and year else "leave_applications.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        
        return response

